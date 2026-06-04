from __future__ import annotations

import ipaddress
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


SOURCE = Path(r"C:\Users\QC_PA\Downloads\IED NAME & IP Address_GI Cikande New_150kV_A1 1.xlsx")
OUTPUT = Path("outputs/ARNetDiscovery_Cikande_Target_Import.xlsx")


def clean(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def find_header_row(raw: pd.DataFrame) -> int:
    for index, row in raw.iterrows():
        values = {clean(v).upper() for v in row.tolist()}
        if "IP ADDRESS" in values and ("NAMA IED" in values or "JENIS PERALATAN" in values):
            return int(index)
    raise RuntimeError("Could not find the IP target table header.")


def is_ipv4(value: str) -> bool:
    try:
        ipaddress.IPv4Address(value)
        return value.count(".") == 3
    except ValueError:
        return False


def load_targets() -> list[dict[str, str]]:
    raw = pd.read_excel(SOURCE, sheet_name="IP", header=None, dtype=str)
    header_row = find_header_row(raw)
    data = pd.read_excel(SOURCE, sheet_name="IP", header=header_row, dtype=str)

    rows: list[dict[str, str]] = []
    current_no = ""
    current_bay = ""
    current_bay_type = ""
    for _, row in data.iterrows():
        ip = clean(row.get("IP ADDRESS"))
        if not is_ipv4(ip):
            continue

        no = clean(row.get("NO"))
        bay = clean(row.get("NAMA BAY"))
        bay_type = clean(row.get("JENIS BAY"))
        if no:
            current_no = no
        if bay:
            current_bay = bay
        if bay_type:
            current_bay_type = bay_type

        expected_type = clean(row.get("JENIS PERALATAN"))
        rows.append(
            {
                "IP ADDRESS": ip,
                "NAMA IED": clean(row.get("NAMA IED")) or expected_type or ip,
                "JENIS PERALATAN": expected_type,
                "PANEL": clean(row.get("PANEL")),
                "NO DEVICE": clean(row.get("NO DEVICE")) or current_no,
                "NAMA BAY": bay or current_bay,
                "JENIS BAY": bay_type or current_bay_type,
                "REMARK": clean(row.get("REMARK")),
                "SOURCE": "IED NAME & IP ADDRESS 150kV GI CIKANDE NEW / IP",
            }
        )

    ntp_rows = [
        {
            "IP ADDRESS": "1.108.200.241",
            "NAMA IED": "NTP SERVER 1",
            "JENIS PERALATAN": "NTP Server",
            "PANEL": "",
            "NO DEVICE": "",
            "NAMA BAY": "Infrastructure",
            "JENIS BAY": "Server",
            "REMARK": "Detected above IP target table",
            "SOURCE": "IED NAME & IP ADDRESS 150kV GI CIKANDE NEW / IP",
        },
        {
            "IP ADDRESS": "1.108.200.242",
            "NAMA IED": "NTP SERVER 2",
            "JENIS PERALATAN": "NTP Server",
            "PANEL": "",
            "NO DEVICE": "",
            "NAMA BAY": "Infrastructure",
            "JENIS BAY": "Server",
            "REMARK": "Detected above IP target table",
            "SOURCE": "IED NAME & IP ADDRESS 150kV GI CIKANDE NEW / IP",
        },
    ]

    unique: dict[str, dict[str, str]] = {}
    for row in ntp_rows + rows:
        unique.setdefault(row["IP ADDRESS"], row)
    return sorted(unique.values(), key=lambda r: tuple(int(p) for p in r["IP ADDRESS"].split(".")))


def style_sheet(ws, row_count: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="0B1020")
    header_font = Font(color="FFFFFF", bold=True, name="Segoe UI")
    body_font = Font(color="172033", name="Segoe UI", size=10)
    muted_font = Font(color="5B6B84", name="Segoe UI", size=10)
    thin = Side(style="thin", color="D8E0EA")
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=row_count + 1):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            cell.number_format = "@"
        row[8].font = muted_font

    widths = [16, 24, 28, 12, 12, 28, 18, 20, 46]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    table_ref = f"A1:I{row_count + 1}"
    table = Table(displayName="ARNetTargets", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_workbook() -> None:
    targets = load_targets()
    wb = Workbook()
    ws = wb.active
    ws.title = "ARNet Import"
    headers = [
        "IP ADDRESS",
        "NAMA IED",
        "JENIS PERALATAN",
        "PANEL",
        "NO DEVICE",
        "NAMA BAY",
        "JENIS BAY",
        "REMARK",
        "SOURCE",
    ]
    ws.append(headers)
    for row in targets:
        ws.append([row[h] for h in headers])
    style_sheet(ws, len(targets))

    readme = wb.create_sheet("README")
    readme.sheet_view.showGridLines = False
    readme["A1"] = "ARNet Discovery Import"
    readme["A1"].font = Font(name="Segoe UI", bold=True, size=16, color="0B1020")
    readme["A3"] = "Use the 'ARNet Import' sheet directly with the application's Import List button."
    readme["A4"] = "The app reads IP ADDRESS plus optional expected device metadata, then Scan List probes exact relay/server targets."
    readme["A6"] = f"Targets exported: {len(targets)}"
    for cell in readme["A"]:
        font = copy(cell.font)
        font.name = "Segoe UI"
        cell.font = font
    readme.column_dimensions["A"].width = 110

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT.resolve())
    print(f"targets={len(targets)}")


if __name__ == "__main__":
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)
    build_workbook()
